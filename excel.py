from openpyxl import load_workbook
wb = load_workbook('化學成績登記.xlsx')

ws = wb.active

while(True):
    order = input("enter? plus? initial? \n")   
    if order == "initial":  #initial excel  
        for i in range(1,40):
            ws['A' + str(i)] = 0

    elif order == "close":  #end program
        break
    
    while order == "enter": #enter the score
        number = input("號碼: ")
        grades = input("分數: ")
        if number == "end" and grades == "end":
            break
        else:
            ws['A'+ number] = int(grades)
            print('\n')
            
    if order == "plus": #plus score to the number
        score = int(input("要加的分數: "))
        while(True):
            plusnumber = input("要加分的人: ")
            if(plusnumber == "end"):
                break
            else:
                ws['A' + plusnumber].value += score

                


wb.save('化學成績登記.xlsx')

